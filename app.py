import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


# -----------------------------
# Helpers
# -----------------------------
def clean_name(value) -> str:
    """Return a cleaned name string or empty string."""
    if pd.isna(value) or str(value).strip() == "":
        return ""
    s = str(value).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def normalize_key(value: str) -> str:
    """Normalize an identifier for matching (email or name)."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def standardize_columns(temp_df: pd.DataFrame, allow_name_fallback: bool, filename: str):
    """
    Normalizes columns from Google Forms / Microsoft Forms exports.

    Required:
      - "Start time" (Microsoft) OR "Timestamp" (Google)

    Optional:
      - "Email" OR "Email Address"
      - One of several name columns (Name, Full Name, etc.)

    Returns:
      standardized_df with columns:
        - Start time
        - Email
        - DisplayName
        - StudentKey (internal)
      report dict for diagnostics
    """
    df = temp_df.copy()
    original_cols = list(df.columns)

    # --- Time column ---
    detected_time = None
    if "Start time" in df.columns:
        detected_time = "Start time"
    elif "Timestamp" in df.columns:
        df = df.rename(columns={"Timestamp": "Start time"})
        detected_time = "Timestamp"

    if "Start time" not in df.columns:
        raise ValueError(
            f"[{filename}] Missing timestamp column. Expected 'Start time' (Microsoft) or 'Timestamp' (Google)."
        )

    # --- Email column ---
    detected_email = None
    if "Email" in df.columns:
        detected_email = "Email"
    elif "Email Address" in df.columns:
        df = df.rename(columns={"Email Address": "Email"})
        detected_email = "Email Address"

    # --- Name column(s) ---
    name_candidates = [
        "Full Name",
        "Name",
        "Student Name",
        "Respondent",
        "Name (First Last)",
        "Your Name",
        "Name1",
        "Name2",
    ]
    detected_name = next((c for c in name_candidates if c in df.columns), None)

    if detected_name:
        df["DisplayName"] = df[detected_name].apply(clean_name)
    else:
        df["DisplayName"] = ""

    # --- Build StudentKey ---
    used_identifier = None

    if "Email" in df.columns:
        df["Email"] = df["Email"].astype(str).str.strip().str.lower()
        df.loc[df["Email"].isin(["nan", "none"]), "Email"] = ""
        # Prefer email; if email blank and fallback enabled, use name
        if allow_name_fallback:
            df["StudentKey"] = df.apply(
                lambda r: normalize_key(r["Email"]) if normalize_key(r["Email"]) else normalize_key(r["DisplayName"]),
                axis=1,
            )
            used_identifier = "Email (fallback to Name if blank)"
        else:
            df["StudentKey"] = df["Email"].apply(normalize_key)
            used_identifier = "Email"
    else:
        # No email column in file
        df["Email"] = ""
        if not allow_name_fallback:
            raise ValueError(
                f"[{filename}] Missing email column. Expected 'Email' (Microsoft) or 'Email Address' (Google). "
                f"Enable 'Allow name fallback' to proceed without email."
            )
        df["StudentKey"] = df["DisplayName"].apply(normalize_key)
        used_identifier = "Name (fallback; no Email column)"

    # Validate StudentKey not blank
    if (df["StudentKey"].astype(str).str.len() == 0).any():
        raise ValueError(
            f"[{filename}] Some rows have no usable identifier. Ensure Email is collected or names are filled in."
        )

    report = {
        "File": filename,
        "Detected Time Column": detected_time or "NOT FOUND",
        "Detected Email Column": detected_email or "NOT FOUND",
        "Detected Name Column": detected_name or "NOT FOUND",
        "Identifier Used": used_identifier,
        "Columns in File": ", ".join(original_cols[:12]) + ("..." if len(original_cols) > 12 else ""),
    }

    return df[["Start time", "Email", "DisplayName", "StudentKey"]], report


def format_workbook_in_memory(output_bytes: BytesIO) -> bytes:
    output_bytes.seek(0)
    wb = load_workbook(output_bytes)
    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Header row
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = "A2"

        # Column widths (cap work to first 200 rows)
        for col in ws.columns:
            col_letter = col[0].column_letter
            max_len = 0
            for cell in col[:200]:
                if cell.value is not None and str(cell.value) != "":
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

    final_output = BytesIO()
    wb.save(final_output)
    return final_output.getvalue()


# -----------------------------
# Streamlit App
# -----------------------------
st.set_page_config(page_title="Attendance Processor", layout="centered")
st.title("📊 Attendance Processor")

allow_name_fallback = st.checkbox(
    "Allow name fallback if Email is missing/blank (less reliable)",
    value=False,
    help="If Email is missing or blank, the app will use a normalized full name as the identifier. "
         "This can miscount students with the same name or inconsistent spelling.",
)

include_diagnostics_sheet = st.checkbox(
    "Include a Diagnostics sheet (file column detection report)",
    value=True,
)

uploaded_files = st.file_uploader(
    "Upload Attendance Files (.xlsx)",
    type="xlsx",
    accept_multiple_files=True,
)

if uploaded_files and st.button("GENERATE COMPLETE REPORT"):
    try:
        all_dfs = []
        file_reports = []

        for f in uploaded_files:
            temp_df = pd.read_excel(f)
            standardized_df, report = standardize_columns(
                temp_df=temp_df,
                allow_name_fallback=allow_name_fallback,
                filename=getattr(f, "name", "uploaded_file.xlsx"),
            )
            all_dfs.append(standardized_df)
            file_reports.append(report)

        # Show detection results
        st.subheader("Detected Columns / Mapping (per file)")
        diag_df = pd.DataFrame(file_reports)
        st.dataframe(diag_df, use_container_width=True)

        # Combine
        df = pd.concat(all_dfs, ignore_index=True)

        # Parse time + derive date
        df["SubmissionDateTime_ET"] = pd.to_datetime(df["Start time"], errors="coerce")
        df = df.dropna(subset=["SubmissionDateTime_ET"]).copy()
        df["ClassDate"] = df["SubmissionDateTime_ET"].dt.normalize()

        # Normalize fields
        df["StudentKey"] = df["StudentKey"].astype(str).str.strip().str.lower()
        df["Email"] = df["Email"].astype(str).str.strip().str.lower()
        df["DisplayName"] = df["DisplayName"].apply(clean_name)

        if df.empty:
            raise ValueError("No valid attendance records found after parsing timestamps.")

        # Deduplicate: first entry per student per day (by StudentKey)
        daily = (
            df.sort_values(["ClassDate", "StudentKey", "SubmissionDateTime_ET"])
              .drop_duplicates(subset=["ClassDate", "StudentKey"], keep="first")
              .copy()
        )

        total_days = int(daily["ClassDate"].nunique())
        if total_days == 0:
            raise ValueError("No class days detected after deduplication.")

        # Stable name/email mapping per StudentKey
        def mode_nonblank(series: pd.Series, default: str):
            s = series.fillna("").astype(str)
            s = s[s.str.strip() != ""]
            return s.value_counts().idxmax() if not s.empty else default

        name_map = daily.groupby("StudentKey")["DisplayName"].agg(lambda s: mode_nonblank(s, "Unknown"))
        email_map = daily.groupby("StudentKey")["Email"].agg(lambda s: mode_nonblank(s, ""))

        # -----------------------------
        # Sheets (match your original template layout)
        # -----------------------------
        # Student_Summary: DisplayName, Email, DaysPresent, TotalClassDays, AttendancePercent, LastSeen
        student_summary = (
            daily.groupby("StudentKey")
                 .agg(DaysPresent=("ClassDate", "nunique"), LastSeen=("ClassDate", "max"))
                 .reset_index()
        )
        student_summary["DisplayName"] = student_summary["StudentKey"].map(name_map)
        student_summary["Email"] = student_summary["StudentKey"].map(email_map)
        student_summary["TotalClassDays"] = total_days
        student_summary["AttendancePercent"] = (student_summary["DaysPresent"] / total_days * 100).round(1)
        student_summary = student_summary[["DisplayName", "Email", "DaysPresent", "TotalClassDays", "AttendancePercent", "LastSeen"]]

        # Daily_Attendance: ClassDate, Email, DisplayName, SubmissionDateTime_ET
        daily_attendance = daily[["ClassDate", "StudentKey", "SubmissionDateTime_ET"]].copy()
        daily_attendance["DisplayName"] = daily_attendance["StudentKey"].map(name_map)
        daily_attendance["Email"] = daily_attendance["StudentKey"].map(email_map)
        daily_attendance = daily_attendance[["ClassDate", "Email", "DisplayName", "SubmissionDateTime_ET"]]

        # Per_Day_Counts: ClassDate, PresentCount
        per_day = daily.groupby("ClassDate").agg(PresentCount=("StudentKey", "nunique")).reset_index()

        # Most_Recent_By_Student: ClassDate, Email, DisplayName, SubmissionDateTime_ET
        most_recent = (
            daily.sort_values("ClassDate")
                 .drop_duplicates(subset="StudentKey", keep="last")
                 .copy()
        )
        most_recent["DisplayName"] = most_recent["StudentKey"].map(name_map)
        most_recent["Email"] = most_recent["StudentKey"].map(email_map)
        most_recent_sheet = most_recent[["ClassDate", "Email", "DisplayName", "SubmissionDateTime_ET"]].copy()

        # Student_Status_Report: ClassDate, Email, DisplayName, SubmissionDateTime_ET, DaysPresent, TotalClassDays, AttendancePercent
        status_report = most_recent_sheet.merge(
            student_summary[["Email", "DaysPresent", "TotalClassDays", "AttendancePercent"]],
            on="Email",
            how="left",
        ).sort_values("ClassDate", ascending=False)

        status_report = status_report[
            ["ClassDate", "Email", "DisplayName", "SubmissionDateTime_ET", "DaysPresent", "TotalClassDays", "AttendancePercent"]
        ]

        # -----------------------------
        # Export Excel
        # -----------------------------
        excel_data = BytesIO()
        with pd.ExcelWriter(excel_data, engine="openpyxl") as writer:
            student_summary.to_excel(writer, sheet_name="Student_Summary", index=False)
            daily_attendance.to_excel(writer, sheet_name="Daily_Attendance", index=False)
            per_day.to_excel(writer, sheet_name="Per_Day_Counts", index=False)
            most_recent_sheet.to_excel(writer, sheet_name="Most_Recent_By_Student", index=False)
            status_report.to_excel(writer, sheet_name="Student_Status_Report", index=False)
            if include_diagnostics_sheet:
                diag_df.to_excel(writer, sheet_name="Diagnostics", index=False)

        formatted_excel = format_workbook_in_memory(excel_data)

        if allow_name_fallback:
            st.warning(
                "Name fallback is enabled. If two students share the same name (or names are inconsistent), counts may be incorrect. "
                "Collecting emails is strongly recommended."
            )

        st.success("✅ Clean Report Generated matching your template!")
        st.download_button(
            label="📥 Download Excel Report",
            data=formatted_excel,
            file_name="Attendance_Summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error(f"Error processing files: {e}")
